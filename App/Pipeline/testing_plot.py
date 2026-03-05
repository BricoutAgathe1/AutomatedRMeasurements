# python
import math
import os
import numpy as np
import matplotlib.pyplot as plt
import openpyxl


def plot_pipe_lengths(xlsx_path="results.xlsx", show=True):
    # Read pipe lengths
    if not os.path.isfile(xlsx_path):
        print(f"`{xlsx_path}` not found.")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "PipeLengths" not in wb.sheetnames:
        print("Sheet `PipeLengths` not found.")
        return
    ws = wb["PipeLengths"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # skip header
    if not rows:
        print("No data in `PipeLengths`.")
        return

    diameters = np.array([float(d) for d, l in rows])
    lengths = np.array([float(l) for d, l in rows])

    # Vectorized inverse-diameter computation (replaces the broken for-loop)
    # original per-point formula: inv_d = 2 / sqrt(d*d / cos(40deg))
    cos40 = math.cos(math.radians(40))
    inv_d = 2.0 / np.sqrt(diameters ** 2 / cos40)

    order = np.argsort(inv_d)
    x = inv_d[order]
    y = lengths[order]

    lastSeen = 0.7
    x_end = 2 / math.sqrt(lastSeen * lastSeen / cos40)

    # Build augmented dataset: include measured points + (x_end, 0)
    x_aug = np.concatenate((x, [x_end]))
    y_aug = np.concatenate((y, [0.0]))

    # sort augmented arrays by x (necessary for interpolation/trapz)
    idx = np.argsort(x_aug)
    x_aug = x_aug[idx]
    y_aug = y_aug[idx]

    # Interpolate on a fine grid from 0 to x_end (ensure xi includes x_end)
    xi = np.linspace(0.0, x_end, 800)
    yi = np.interp(xi, x_aug, y_aug)

    # Compute AUC from x=0 to x=x_end using interpolated curve
    auc = float(np.trapz(yi, xi))
    r = auc/2

    # Cumulative trapezoid by segment (robust, no SciPy needed)
    dx = np.diff(xi)
    seg_area = 0.5 * (yi[:-1] + yi[1:]) * dx
    cum = np.concatenate(([0.0], np.cumsum(seg_area)))  # same length as x

    target = auc * 0.5
    # find first index where cumulative area >= target
    idx = np.searchsorted(cum, target, side="left")
    if idx == 0:
        alpha_R = xi[0]
    else:
        # linear interpolate between x[idx-1] and x[idx]
        a0, a1 = xi[idx - 1], xi[idx]
        c0, c1 = cum[idx - 1], cum[idx]
        if c1 == c0:
            alpha_R = a0
        else:
            frac = (target - c0) / (c1 - c0)
            alpha_R = a0 + frac * (a1 - a0)

    # L_R is the curve value at alpha_R (linear interpolation)
    L_R = float(np.interp(alpha_R, xi, yi))
    alpha_R = float(auc / L_R) if L_R > 0 else np.nan
    D_R = 2 / alpha_R

    # Use pyplot-managed figure so plt.show() will display it
    # Plot
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(x, y, 'o', color='tab:blue', label='Measured')
    ax.plot(xi, yi, '-', color='tab:blue', alpha=0.9, label=f'Interpolated')

    ax.plot(0,L_R, 'gP', zorder=10, clip_on=False, label=f'L_R = {L_R:.0f} mm')
    ax.plot(alpha_R,0, 'rx', zorder=10, clip_on=False, label=f'alpha_R = {alpha_R:.2f} mm⁻¹')

    # axes start at 0,0 and set limits a bit beyond values for readability
    x_max = x_end
    y_max = y.max()
    ax.set_xlim(0, x_max * 1.05)
    ax.set_ylim(0, y_max * 1.10)

    ax.set_xlabel("alpha (mm⁻¹)")
    ax.set_ylabel("L (mm)")
    ax.set_title(f"L vs alpha  (R={r:.0f})")
    ax.grid(True)
    ax.legend()

    # Print values for quick inspection
    print(f"R = {r:.0f}")
    print(f"LR = {L_R:.0f}")
    print(f"DR = {D_R:.2f}")

    if show:
        plt.show()
    return {
        "auc": auc,
        "x": x,
        "y": y,
        "xi": xi,
        "yi": yi
    }


def main():
    plot_pipe_lengths(xlsx_path="../results.xlsx", show=True)


if __name__ == "__main__":
    main()

