import sys
import os
import cv2
import json
from PySide6 import QtWidgets, QtGui, QtCore
from UI.ui_form import Ui_MainWindow
from Pipeline.library import extract_frames, checkbox_style, button_style, progress_style, QWidget_style, textbox_style, calculate_pipe_lengths
from multiprocessing import Pool
import segmentation_models_pytorch as smp
import torch
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import math


class ImageLabel(QtWidgets.QLabel):
    squareDrawn = QtCore.Signal(QtCore.QRect)
    lineDrawn = QtCore.Signal(tuple)  # (start, end)

    def __init__(self, main_window, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.main_window = main_window # store reference
        self.start_pos = None
        self.end_pos = None
        self.drawing_mode = "rect"  # "rect" or "line"
        self.rect = None
        self.line = None
        self._pixmap = None
        self._pixmap_draw_rect = QtCore.QRect()  # area where pixmap is drawn inside the label
        self.force_vertical = True
        self.setMouseTracking(True)
        self.setAlignment(QtCore.Qt.AlignCenter)

    def setPixmap(self, pixmap):
        self._pixmap = pixmap
        super().setPixmap(pixmap)

    def set_mode(self, mode):
        self.drawing_mode = mode
        self.start_pos = None
        self.end_pos = None
        self.update()

    def mousePressEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton:
            self.start_pos = event.position().toPoint()
            self.end_pos = event.position().toPoint()
            self.update()

    def mouseMoveEvent(self, event):
        if self.start_pos:
            if self.drawing_mode == "rect":
                # force square
                dx = event.position().x() - self.start_pos.x()
                dy = event.position().y() - self.start_pos.y()
                side = max(abs(dx), abs(dy))
                self.end_pos = QtCore.QPoint(
                    self.start_pos.x() + (side if dx >= 0 else -side),
                    self.start_pos.y() + (side if dy >= 0 else -side)
                )
            elif self.drawing_mode == "line":
                if self.main_window.checkVertical.isChecked():
                    self.end_pos = QtCore.QPoint(self.start_pos.x(), event.position().toPoint().y())
                else:
                    self.end_pos = event.position().toPoint()
            self.update()

    def mouseReleaseEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton and self.start_pos and self.end_pos:
            if self.drawing_mode == "rect":
                self.rect = QtCore.QRect(self.start_pos, self.end_pos).normalized()
                self.squareDrawn.emit(self.rect)
            elif self.drawing_mode == "line":
                self.line = (self.start_pos, self.end_pos)
                self.lineDrawn.emit(self.line)
            self.start_pos = None
            self.end_pos = None
            self.update()

    def paintEvent(self, event):
        painter = QtGui.QPainter(self)


        # draw image (centered, keep aspect ratio) and remember draw rect
        if self._pixmap and not self._pixmap.isNull():
            scaled = self._pixmap.scaled(
                self.contentsRect().size(),
                QtCore.Qt.KeepAspectRatio,
                QtCore.Qt.SmoothTransformation
            )
            cr = self.contentsRect()
            x = cr.x() + (cr.width() - scaled.width()) // 2
            y = cr.y() + (cr.height() - scaled.height()) // 2
            self._pixmap_draw_rect = QtCore.QRect(x, y, scaled.width(), scaled.height())
            painter.drawPixmap(self._pixmap_draw_rect, scaled)
        else:
            self._pixmap_draw_rect = QtCore.QRect()

        # draw overlays
        pen = QtGui.QPen(QtGui.QColor("red"), 2, QtCore.Qt.SolidLine)
        painter.setPen(pen)

        if self.drawing_mode == "rect" and (self.start_pos and self.end_pos):
            painter.drawRect(QtCore.QRect(self.start_pos, self.end_pos).normalized())
        elif self.drawing_mode == "rect" and self.rect:
            painter.drawRect(self.rect)

        elif self.drawing_mode == "line" and (self.start_pos and self.end_pos):
            painter.drawLine(self.start_pos, self.end_pos)
        elif self.drawing_mode == "line" and self.line:
            painter.drawLine(self.line[0], self.line[1])

    def widget_rect_to_pixmap_rect(self, widget_rect: QtCore.QRect) -> QtCore.QRect:
        """
        Map a QRect in widget coordinates (e.g. the user-drawn rect) to pixmap pixel coordinates.
        Returns a QRect clipped to pixmap bounds, or None if mapping is not possible.
        """
        if not self._pixmap or self._pixmap_draw_rect.isNull():
            return None

        # local coords inside drawn pixmap
        local_x = widget_rect.x() - self._pixmap_draw_rect.x()
        local_y = widget_rect.y() - self._pixmap_draw_rect.y()
        local_w = widget_rect.width()
        local_h = widget_rect.height()

        # If the user's rect is outside the drawn pixmap, clip it
        # compute scale factors from draw rect -> actual pixmap pixels
        scale_w = self._pixmap.width() / self._pixmap_draw_rect.width()
        scale_h = self._pixmap.height() / self._pixmap_draw_rect.height()
        px = int(local_x * scale_w)
        py = int(local_y * scale_h)
        pw = int(local_w * scale_w)
        ph = int(local_h * scale_h)

        mapped = QtCore.QRect(px, py, pw, ph).normalized()
        mapped = mapped.intersected(QtCore.QRect(0, 0, self._pixmap.width(), self._pixmap.height()))
        return mapped if not mapped.isNull() else None

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.imageLabel = ImageLabel(self)

        self.btnSelectFolder.setStyleSheet(button_style)
        self.btnSetCrop.setStyleSheet(button_style)
        self.btnSaveCalibration.setStyleSheet(button_style)
        self.btnNextImage.setStyleSheet(button_style)
        self.btnBatchCrop.setStyleSheet(button_style)
        self.btnRunBatch.setStyleSheet(button_style)
        self.progressBarVideo.setStyleSheet(progress_style)
        self.progressBar.setStyleSheet(progress_style)
        self.progressBar_2.setStyleSheet(progress_style)

        self.btnChecks.setStyleSheet(checkbox_style)
        self.btnAmbientLight.setStyleSheet(checkbox_style)
        self.btnDisplayLight.setStyleSheet(checkbox_style)
        self.btnGenerate.setStyleSheet(button_style)
        self.txtComments.setStyleSheet(textbox_style)
        self.widget_12.setStyleSheet(QWidget_style)

        self.checkbox_group = QtWidgets.QButtonGroup(self)
        self.checkbox_group.setExclusive(True)
        self.checkbox_group.addButton(self.checkVertical)
        self.checkbox_group.addButton(self.checkOther)

        self.checkVertical.setStyleSheet(checkbox_style)
        self.checkOther.setStyleSheet(checkbox_style)
        self.checkVertical.setChecked(True)  # Default selection

        self.checkVertical.stateChanged.connect(self.on_vertical_checked)
        self.checkOther.stateChanged.connect(self.on_other_checked)

        self.imageLabel = ImageLabel(main_window=self, parent=self.lblImageDisplay.parent())
        self.imageLabel.setGeometry(self.lblImageDisplay.geometry())  # same size & position
        self.imageLabel.setPixmap(self.lblImageDisplay.pixmap())
        self.imageLabel.setScaledContents(self.lblImageDisplay.hasScaledContents())

        self.lblImageDisplay.hide()  # keep formatting, just hide it

        # Connect drawing signals
        self.imageLabel.squareDrawn.connect(self.on_square_drawn)
        self.imageLabel.lineDrawn.connect(self.on_line_drawn)

        self.show_image("graphics/Thesis - graphics.png")

        self.btnSelectFolder.clicked.connect(self.select_folder)
        self.btnSetCrop.clicked.connect(self.on_set_crop)
        self.btnSaveCalibration.clicked.connect(self.on_save_calibration)
        self.btnNextImage.setEnabled(False)
        self.btnNextImage.clicked.connect(self.next_image)
        self.btnRunBatch.setEnabled(False)
        self.btnBatchCrop.setEnabled(False)
        self.btnBatchCrop.clicked.connect(self.batch_crop_all_folders)
        self.progressBar.setEnabled(False)
        self.progressBar_2.setEnabled(False)
        self.progressBarVideo.setEnabled(False)
        self.btnRunBatch.clicked.connect(self.batch_segment_all_folders)

        self.rect = None
        self.line = None

    def on_vertical_checked(self, checked):
        if checked:
            self.imageLabel.force_vertical = True

    def on_other_checked(self, checked):
        if checked:
            self.imageLabel.force_vertical = False

    def show_image(self, image_path):
        """Display an image in the image label."""
        pixmap = QtGui.QPixmap(image_path)
        self.lblImageGuideCrop.setPixmap(pixmap)
        self.lblImageGuideCrop.setScaledContents(True)

    def select_folder(self):
        folder_path = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Data Folder")
        self.progressBarVideo.setEnabled(True)

        if folder_path:
            scannerName = self.txtScannerName.text().strip()
            self.lblScannerNameResults.setText(str(scannerName))

            self.lblStatus.setText("Status: Extracting frames...")

            self.lblFolderPath.setText(folder_path)
            self.video_folders = [os.path.join(folder_path, f)
                                  for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f))]
            self.current_video_idx = 0
            self.cropping_coordinates = {}
            self.conversion_factors = {}

            self.worker = FrameExtractorWorker(folder_path, extract_frames)
            self.worker.progress.connect(self.progressBarVideo.setValue)  # Link progress bar
            self.worker.finished.connect(self.load_first_image_for_cropping)
            self.worker.start()

    def on_square_drawn(self, rect):
        self.rect = rect

    def on_line_drawn(self, line):
        self.line = line

    def update_next_button_state(self):
        self.btnNextImage.setEnabled(self.rect is not None and self.line is not None)

    def load_first_image_for_cropping(self):
        # If video_folders not set or empty, (re)discover subfolders in the selected base folder
        if not getattr(self, "video_folders", None) or len(self.video_folders) == 0:
            base = self.lblFolderPath.text()
            if not base or not os.path.isdir(base):
                self.lblStatus.setText("Status: No folders found")
                return
            self.video_folders = [
                os.path.join(base, f)
                for f in os.listdir(base)
                if os.path.isdir(os.path.join(base, f))
            ]
            self.current_video_idx = 0
            self.cropping_coordinates = getattr(self, "cropping_coordinates", {})
            self.conversion_factors = getattr(self, "conversion_factors", {})

        if self.current_video_idx >= len(self.video_folders):
            self.btnNextImage.setEnabled(False)
            self.btnBatchCrop.setEnabled(True)
            return

        current_folder = self.video_folders[self.current_video_idx]
        self.btnSelectFolder.setEnabled(False)

        # Update folder number label
        self.lblFolderNumber.setText(f"{self.current_video_idx + 1} / {len(self.video_folders)}")
        self.lblStatus.setText("Status: Editing frames...")

        frames_dir = os.path.join(current_folder, "frames")
        # guard against missing/empty frames directory
        if not os.path.isdir(frames_dir):
            self.lblStatus.setText(f"Status: No frames in {current_folder}")
            return
        entries = sorted(os.listdir(frames_dir))
        if not entries:
            self.lblStatus.setText(f"Status: No frames in {frames_dir}")
            return

        first_image_path = os.path.join(frames_dir, entries[0])

        self.current_image_path = first_image_path
        self.rect = None
        self.line = None
        self.update_next_button_state()

        # Load full (original) image into ImageLabel so the label can scale it consistently
        pixmap = QtGui.QPixmap(first_image_path)
        if pixmap.isNull():
            self.lblStatus.setText("Status: Failed to load image")
            return

        self.imageLabel.setPixmap(pixmap)
        self.imageLabel.setScaledContents(False)

        # Reset drawing stuff
        self.imageLabel.rect = None
        self.imageLabel.line = None

    # python
    def on_set_crop(self):
        if not self.imageLabel.rect:
            return

        # Map the widget-space rect to pixmap (original image) pixels
        mapped = self.imageLabel.widget_rect_to_pixmap_rect(self.imageLabel.rect)
        if mapped is None:
            return

        # Store rect in original image pixel coordinates so BatchCropWorker can reuse it
        self.cropping_coordinates[self.current_video_idx] = mapped

        # Use original pixmap to create a correct preview crop
        orig_pm = QtGui.QPixmap(self.current_image_path)
        if orig_pm.isNull():
            return

        cropped = orig_pm.copy(mapped)
        final = cropped.scaled(
            256, 256,
            QtCore.Qt.IgnoreAspectRatio,
            QtCore.Qt.SmoothTransformation
        )

        self.imageLabel.setScaledContents(False)
        self.imageLabel.setPixmap(final)
        self.imageLabel.setFixedSize(final.size())
        self.imageLabel.setAlignment(QtCore.Qt.AlignCenter)

        self.imageLabel.drawing_mode = "line"
        self.imageLabel.line = None

        self.frameConversion.setEnabled(True)

    def on_save_calibration(self):
        """Save calibration line + distance."""

        if not self.imageLabel.line:
            return

        dist_mm_str = self.txtDistanceMm.text().strip()
        if not dist_mm_str:
            return

        try:
            dist_mm = float(dist_mm_str)
        except ValueError:
            return

        # Extract coordinates
        start_point, end_point = self.imageLabel.line
        x1, y1 = start_point.x(), start_point.y()
        x2, y2 = end_point.x(), end_point.y()

        dx = x2 - x1
        dy = y2 - y1

        # Compute total pixel distance along line
        line_length_px = (dx ** 2 + dy ** 2) ** 0.5

        if line_length_px == 0:
            return

        if self.checkVertical.isChecked():
            # Vertical line: full pixel distance corresponds to mm
            vertical_distance_px = abs(dy)
            conversion_factor = dist_mm / vertical_distance_px
        else:
            # Angled line: project user distance onto vertical axis
            vertical_distance_px = abs(dy)
            vertical_distance_mm = dist_mm * (vertical_distance_px / line_length_px)
            conversion_factor = vertical_distance_mm / vertical_distance_px

        # Store result
        self.conversion_factors[self.current_video_idx] = conversion_factor

        self.lblStatus.setText(f"Calibration saved: {conversion_factor:.4f} mm/pixel")
        self.btnNextImage.setEnabled(True)

        self.btnNextImage.setEnabled(True)

    def next_image(self):
        """Go to next folder or finish all."""
        self.current_video_idx += 1
        if self.current_video_idx >= len(self.video_folders):
            # last folder → disable buttons and batch crop
            self.btnNextImage.setEnabled(False)
            self.btnBatchCrop.setEnabled(True)
            self.btnSaveCalibration.setEnabled(True)
            self.btnSetCrop.setEnabled(True)

            # reset label back to full display mode
            self.imageLabel.setScaledContents(True)
            self.imageLabel.setFixedSize(self.lblImageDisplay.size())  # reset to container size
            self.imageLabel.setAlignment(QtCore.Qt.AlignCenter)
            pixmap = QtGui.QPixmap("graphics/blank.png")
            self.imageLabel.line=None
            self.imageLabel.rect = None
            self.imageLabel.setPixmap(pixmap)
            return

        # reset drawing state
        self.imageLabel.drawing_mode = "rect"
        self.imageLabel.start_pos = None
        self.imageLabel.end_pos = None
        self.imageLabel.rect = None
        self.imageLabel.line = None

        # reset label back to full display mode
        self.imageLabel.setScaledContents(True)
        self.imageLabel.setFixedSize(self.lblImageDisplay.size())  # reset to container size
        self.imageLabel.setAlignment(QtCore.Qt.AlignCenter)

        self.txtDistanceMm.clear()

        # load new first image
        self.load_first_image_for_cropping()

    def batch_crop_all_folders(self):
        self.lblStatus.setText("Status: Batch cropping frames...")
        self.progressBar_2.setEnabled(True)
        self.btnRunBatch.setEnabled(True)
        self.btnSetCrop.setEnabled(False)
        self.btnSaveCalibration.setEnabled(False)
        self.worker = BatchCropWorker(self.video_folders, self.cropping_coordinates)
        self.worker.progress.connect(self.progressBar_2.setValue)
        self.worker.start()

    def batch_segment_all_folders(self):
        self.lblStatus.setText("Status: Batch segmenting images...")
        self.progressBar.setEnabled(True)
        self.btnRunBatch.setEnabled(False)
        self.btnBatchCrop.setEnabled(False)

        # Load model
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        model = smp.Unet(encoder_name="resnet18", in_channels=1, classes=2).to(device)
        model.load_state_dict(torch.load('Model/ChrisScanners_model_weights_unetXresnet18_310725.pth', map_location=torch.device('cpu')))
        model.eval()

        self.worker = BatchSegmentWorker(
            self.video_folders,
            model,
            device,
            self.conversion_factors)

        self.worker.progress.connect(self.progressBar.setValue)
        self.worker.finished.connect(self.export_results)
        self.worker.start()

    def update_excel_with_pipe_lengths(self, xlsx_path, pipe_lengths):
        wb = openpyxl.load_workbook(xlsx_path)

        sheet_name = "PipeLengths"
        # Remove existing sheet if it exists
        if sheet_name in wb.sheetnames:
            std = wb[sheet_name]
            wb.remove(std)

        # Create fresh sheet
        ws = wb.create_sheet(sheet_name)
        ws.append(["Diameter_mm", "PipeLength_mm"])  # headers

        # Insert new rows
        for diameter, length in pipe_lengths:
            ws.append([diameter, length])

        wb.save(xlsx_path)

    def export_results(self):
        xlsx_path = "results.xlsx"
        exporter = ResultsExporter(xlsx_path)

        # Cropping
        crop_rows = [
            [idx, rect.x(), rect.y(), rect.width(), rect.height()]
            for idx, rect in self.cropping_coordinates.items()
        ]
        exporter.add_sheet("Cropping", ["FolderIdx", "X", "Y", "Width", "Height"], crop_rows)

        # Calibration
        calib_rows = [
            [idx, factor] for idx, factor in self.conversion_factors.items()
        ]
        exporter.add_sheet("Calibration", ["FolderIdx", "ConversionFactor"], calib_rows)

        # Segmentation
        seg_rows = [
            [folder,
             self.worker.min_top_distance.get(folder),
             self.worker.max_bottom_distance.get(folder)]
            for folder in self.video_folders
        ]
        exporter.add_sheet("Segmentation", ["Folder", "MinTopDistance_mm", "MaxBottomDistance_mm"], seg_rows)

        # Save first (so file exists before update_excel_with_pipe_lengths runs)
        exporter.save()

        # PipeLengths
        pipe_lengths = calculate_pipe_lengths(self.lblFolderPath.text())  # data_dir
        self.update_excel_with_pipe_lengths(xlsx_path, pipe_lengths)

        self.lblStatus.setText(f"Status: Results exported to {xlsx_path}")
        self.plot_pipe_lengths_in_label(self.lblResults, xlsx_path="results.xlsx")

    def plot_pipe_lengths_in_label(self, lbl, xlsx_path="results.xlsx"):
        # Read pipe lengths
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if "PipeLengths" not in wb.sheetnames:
            return
        ws = wb["PipeLengths"]
        rows = list(ws.iter_rows(values_only=True))[1:]  # skip header
        if not rows:
            return

        diameters = np.array([float(d) for d, l in rows])
        lengths = np.array([float(l) for d, l in rows])

        # Vectorized inverse-diameter computation (replaces the broken for-loop)
        # original per-point formula: inv_d = 2 / sqrt(d*d / cos(40deg))
        cos40 = math.cos(math.radians(40))
        inv_d = 2.0 / np.sqrt(diameters ** 2 / cos40)

        order = np.argsort(inv_d)
        x = inv_d[order]
        y = lengths[order]

        lastSeen = 0.7
        x_end = 2 / math.sqrt(lastSeen * lastSeen / cos40)

        # Build augmented dataset: include measured points + (x_end, 0)
        x_aug = np.concatenate((x, [x_end]))
        y_aug = np.concatenate((y, [0.0]))

        # sort augmented arrays by x (necessary for interpolation/trapz)
        idx = np.argsort(x_aug)
        x_aug = x_aug[idx]
        y_aug = y_aug[idx]

        # Interpolate on a fine grid from 0 to x_end (ensure xi includes x_end)
        xi = np.linspace(0.0, x_end, 800)
        yi = np.interp(xi, x_aug, y_aug)

        # Compute AUC from x=0 to x=x_end using interpolated curve
        auc = float(np.trapz(yi, xi))
        r = auc / 2

        # Cumulative trapezoid by segment (robust, no SciPy needed)
        dx = np.diff(xi)
        seg_area = 0.5 * (yi[:-1] + yi[1:]) * dx
        cum = np.concatenate(([0.0], np.cumsum(seg_area)))  # same length as x

        target = auc * 0.5
        # find first index where cumulative area >= target
        idx = np.searchsorted(cum, target, side="left")
        if idx == 0:
            alpha_R = xi[0]
        else:
            # linear interpolate between x[idx-1] and x[idx]
            a0, a1 = xi[idx - 1], xi[idx]
            c0, c1 = cum[idx - 1], cum[idx]
            if c1 == c0:
                alpha_R = a0
            else:
                frac = (target - c0) / (c1 - c0)
                alpha_R = a0 + frac * (a1 - a0)

        # L_R is the curve value at alpha_R (linear interpolation)
        L_R = float(np.interp(alpha_R, xi, yi))
        alpha_R = float(auc / L_R) if L_R > 0 else np.nan
        D_R = 2 / alpha_R

        # Use pyplot-managed figure so plt.show() will display it
        # Plot
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.plot(x, y, 'o', color='tab:blue', label='Measured')
        ax.plot(xi, yi, '-', color='tab:blue', alpha=0.9, label=f'Interpolated')

        ax.plot(0, L_R, 'gP', zorder=10, clip_on=False, label=f'L_R = {L_R:.0f} mm')
        ax.plot(alpha_R, 0, 'rx', zorder=10, clip_on=False, label=f'alpha_R = {alpha_R:.2f} mm⁻¹')

        # axes start at 0,0 and set limits a bit beyond values for readability
        x_max = x_end
        y_max = y.max()
        ax.set_xlim(0, x_max * 1.05)
        ax.set_ylim(0, y_max * 1.10)

        ax.set_xlabel("alpha (mm⁻¹)")
        ax.set_ylabel("L (mm)")
        ax.set_title(f"L vs alpha  (R={r:.0f})")
        ax.grid(True)
        ax.legend()

        # --- Display figure in QLabel ---
        for child in lbl.children():
            child.setParent(None)  # remove previous canvas if exists

        canvas = FigureCanvas(fig)
        canvas.setParent(lbl)
        canvas.setFixedSize(lbl.width(), lbl.height())
        canvas.draw()

        # --- Display results in Results text box ---
        self.lblR.setText(f"{r:.0f}")
        self.lblDR.setText(f"{D_R:.2f} mm")
        self.lblLR.setText(f"{L_R:.0f} mm")


class FrameExtractorWorker(QtCore.QThread):
    progress = QtCore.Signal(int)
    finished = QtCore.Signal()

    def __init__(self, data_dir, extract_frames_func):
        super().__init__()
        self.data_dir = data_dir
        self.extract_frames = extract_frames

    def run(self):
        video_files = [f for f in os.listdir(self.data_dir) if f.endswith(".mp4")]
        video_paths = [os.path.join(self.data_dir, v) for v in video_files]
        frames_dirs = [os.path.join(self.data_dir, os.path.splitext(v)[0], "frames") for v in video_files]
        cropped_dirs = [os.path.join(self.data_dir, os.path.splitext(v)[0], "cropped") for v in video_files]

        with Pool(processes=os.cpu_count()) as pool:
            for i, _ in enumerate(pool.starmap(self.extract_frames, zip(video_paths, frames_dirs)), start=1):
                progress_percent = int((i / len(video_files)) * 100)
                self.progress.emit(progress_percent)

        self.finished.emit()


class BatchCropWorker(QtCore.QThread):
    progress = QtCore.Signal(int)

    def __init__(self, video_folders, cropping_coordinates):
        super().__init__()
        self.video_folders = video_folders
        self.cropping_coordinates = cropping_coordinates

    def run(self):
        total_images = sum(len(os.listdir(os.path.join(f, "frames"))) for f in self.video_folders)
        done = 0

        for idx, folder in enumerate(self.video_folders):
            rect = self.cropping_coordinates.get(idx)
            if not rect:
                continue

            frames_dir = os.path.join(folder, "frames")
            cropped_dir = os.path.join(folder, "cropped")
            os.makedirs(cropped_dir, exist_ok=True)

            for img_name in sorted(os.listdir(frames_dir)):
                img_path = os.path.join(frames_dir, img_name)
                pm = QtGui.QPixmap(img_path)

                cropped = pm.copy(rect)

                save_pixmap = cropped.scaled(256, 256, QtCore.Qt.IgnoreAspectRatio, QtCore.Qt.SmoothTransformation)
                save_pixmap.save(os.path.join(cropped_dir, img_name))

                done += 1
                self.progress.emit(int(done / total_images * 100))


class BatchSegmentWorker(QtCore.QThread):
    progress = QtCore.Signal(int)
    finished = QtCore.Signal()

    def __init__(self, video_folders, model, device, conversion_factors):
        super().__init__()
        self.video_folders = video_folders
        self.model = model
        self.device = device
        self.conversion_factors = conversion_factors  # dict: {folder_index: factor}
        self.max_bottom_distance = {}
        self.min_top_distance = {}

    def run(self):
        total_images = sum(len(os.listdir(os.path.join(f, "cropped"))) for f in self.video_folders)
        done = 0

        for idx, folder in enumerate(self.video_folders):
            cropped_dir = os.path.join(folder, "cropped")
            pseudolabels_dir = os.path.join(folder, "pseudolabels")
            os.makedirs(pseudolabels_dir, exist_ok=True)

            # Get conversion factor for this folder
            conversion_factor = self.conversion_factors.get(idx)
            if conversion_factor is None:
                print(f"⚠️ No conversion factor found for folder {folder}. Skipping.")
                continue

            max_bottom_distance = float('-inf')
            min_top_distance = float('inf')

            for img_name in sorted(os.listdir(cropped_dir)):
                img_path = os.path.join(cropped_dir, img_name)
                image = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
                if image is None:
                    continue

                image = cv2.resize(image, (256, 256))
                image_tensor = torch.tensor(image, dtype=torch.float32).unsqueeze(0).unsqueeze(0) / 255.0
                image_tensor = image_tensor.to(self.device)

                # Run segmentation
                with torch.no_grad():
                    output = self.model(image_tensor)
                    pseudolabel = torch.argmax(output, dim=1).squeeze().cpu().numpy().astype(np.uint8) * 255

                # Save pseudolabel
                pseudolabel_path = os.path.join(pseudolabels_dir, img_name.replace('.png', '_pseudolabel.png'))
                cv2.imwrite(pseudolabel_path, pseudolabel)

                # Calculate top/bottom distances
                mask = cv2.imread(pseudolabel_path, cv2.IMREAD_GRAYSCALE)
                row_sums = np.sum(mask > 0, axis=1)
                non_zero_rows = np.where(row_sums > 0)[0]
                if len(non_zero_rows) == 0:
                    continue

                top_row = non_zero_rows[0]
                bottom_row = non_zero_rows[-1]

                # Convert pixel indices to mm
                top_mm = top_row * conversion_factor
                bottom_mm = bottom_row * conversion_factor

                # Update per-folder stats
                max_bottom_distance = max(max_bottom_distance, bottom_mm)
                min_top_distance = min(min_top_distance, top_mm)

                done += 1
                self.progress.emit(int(done / total_images * 100))

            # Save per-folder JSON file
            result_json_path = os.path.join(folder, "special_distances_unetXresnet18.json")
            result_data = {}
            if min_top_distance < float('inf'):
                result_data["MinTopDistance_mm"] = min_top_distance
            if max_bottom_distance > float('-inf'):
                result_data["MaxBottomDistance_mm"] = max_bottom_distance

            with open(result_json_path, "w") as f:
                json.dump(result_data, f, indent=2)

            # Store results in worker memory too
            self.min_top_distance[folder] = (
                min_top_distance if min_top_distance < float('inf') else None
            )
            self.max_bottom_distance[folder] = (
                max_bottom_distance if max_bottom_distance > float('-inf') else None
            )

        # Emit finished signal when done
        self.finished.emit()


class ResultsExporter:
    def __init__(self, save_path="results.xlsx"):
        self.save_path = save_path
        # Always create a new workbook
        self.wb = openpyxl.Workbook()
        # remove default sheet
        default = self.wb.active
        self.wb.remove(default)

    def add_sheet(self, name, headers, rows):
        ws = self.wb.create_sheet(title=name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        # auto-size columns
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    def save(self):
        self.wb.save(self.save_path)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
